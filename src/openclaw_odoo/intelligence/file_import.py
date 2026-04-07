"""File import/export with auto-detect, column mapping, and relational field resolution."""
import csv
import io
import os
from typing import Any, Optional

from ..errors import OdooClawError

_ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}


def _validate_filepath(filepath: str, mode: str = "read") -> str:
    """Validate and resolve a file path, preventing path traversal.

    Args:
        filepath: The file path to validate.
        mode: 'read' or 'write'.

    Returns:
        The resolved absolute path.

    Raises:
        OdooClawError: If the path is invalid or unsafe.
    """
    if not filepath or not isinstance(filepath, str):
        raise OdooClawError("File path is required")

    # Block path traversal via .. components
    if ".." in os.path.normpath(filepath):
        raise OdooClawError(f"Path traversal not allowed: {filepath}")

    resolved = os.path.realpath(filepath)

    # Reject symlinks
    if os.path.islink(filepath):
        raise OdooClawError(f"Symbolic links not allowed: {filepath}")

    # Validate extension
    _, ext = os.path.splitext(resolved)
    if ext.lower() not in _ALLOWED_EXTENSIONS:
        raise OdooClawError(f"Unsupported file extension: {ext}")

    if mode == "read" and not os.path.isfile(resolved):
        raise OdooClawError(f"File not found: {filepath}")

    return resolved


_MODEL_SIGNATURES: list[tuple[set[str], str]] = [
    # Order matters: more specific first
    ({"product_id", "quantity", "price_unit", "account_id"}, "account.move.line"),
    ({"product_id", "quantity"}, "sale.order.line"),
    ({"job_title", "department_id"}, "hr.employee"),
    ({"work_email", "department_id"}, "hr.employee"),
    ({"job_title", "work_email"}, "hr.employee"),
    ({"list_price", "default_code"}, "product.template"),
    ({"list_price", "categ_id"}, "product.template"),
    ({"name", "list_price"}, "product.template"),
    ({"email", "phone"}, "res.partner"),
    ({"email", "company"}, "res.partner"),
    ({"email", "city"}, "res.partner"),
    ({"phone", "city"}, "res.partner"),
    ({"name", "email"}, "res.partner"),
    ({"name", "phone"}, "res.partner"),
]

_COMMON_ALIASES: dict[str, str] = {
    "company": "parent_id",
    "mobile": "mobile",
    "website": "website",
    "street": "street",
    "city": "city",
    "zip": "zip",
    "country": "country_id",
    "state": "state_id",
    "title": "title",
    "first name": "name",
    "last name": "name",
    "full name": "name",
}

_EXCLUDE_FROM_TEMPLATE = {"id", "create_uid", "create_date", "write_uid", "write_date",
                          "__last_update", "display_name"}
_EXCLUDE_TYPES_TEMPLATE = {"one2many", "many2many", "binary", "image"}


def _default_export_dir() -> str:
    return os.getcwd()


def detect_model(headers: list[str], registry=None) -> Optional[str]:
    """Auto-detect the Odoo model from CSV/Excel column headers.

    Args:
        headers: List of column header strings.
        registry: Optional ModelRegistry for matching custom model signatures.

    Returns:
        Model name string (e.g. 'res.partner'), or None if unrecognized.
    """
    lower_headers = {h.lower().strip() for h in headers}
    # 1. Hardcoded signatures (builtins) first
    for sig_fields, model in _MODEL_SIGNATURES:
        if sig_fields.issubset(lower_headers):
            return model

    # 2. Registry-based signatures for custom models
    if registry is not None:
        try:
            from ..auto_actions import generate_import_signatures
            sigs = generate_import_signatures(registry)
            for sig_labels, model_name in sigs.items():
                if sig_labels.issubset(lower_headers):
                    return model_name
        except Exception:
            pass  # Registry not available, skip

    return None


def map_columns(headers: list[str], model: str, client=None) -> dict[str, str]:
    """Map CSV/Excel headers to Odoo field names using labels, aliases, and fuzzy matching.

    Args:
        headers: Column header strings from the import file.
        model: Target Odoo model name.
        client: Optional OdooClient for fields_get lookup.

    Returns:
        Dict mapping header strings to Odoo field names.
    """
    # Known field names for common models (fallback when no client)
    _KNOWN_FIELDS: dict[str, set[str]] = {
        "res.partner": {"name", "email", "phone", "mobile", "street", "street2",
                        "city", "zip", "country_id", "state_id", "website",
                        "parent_id", "is_company", "vat", "ref", "comment",
                        "function", "title", "lang"},
        "product.template": {"name", "list_price", "default_code", "categ_id",
                             "type", "description", "sale_ok", "purchase_ok",
                             "barcode", "weight", "volume"},
        "sale.order.line": {"product_id", "product_uom_qty", "price_unit", "name",
                            "discount", "order_id"},
        "hr.employee": {"name", "job_title", "department_id", "work_email",
                        "work_phone", "parent_id", "coach_id"},
        "account.move.line": {"product_id", "quantity", "price_unit", "name",
                              "account_id", "move_id"},
    }

    fields_def = {}
    if client:
        fields_def = client.fields_get(model)

    known = _KNOWN_FIELDS.get(model, set())
    # Build label->field mapping from fields_get
    label_map: dict[str, str] = {}
    for fname, fdef in fields_def.items():
        label = fdef.get("string", "")
        if label:
            label_map[label.lower()] = fname

    result: dict[str, str] = {}
    for header in headers:
        h_lower = header.lower().strip()

        # 1. Direct field name match
        if h_lower in known or h_lower in fields_def:
            result[header] = h_lower
            continue

        # 2. Exact label match from fields_get
        if h_lower in label_map:
            result[header] = label_map[h_lower]
            continue

        # 3. Common alias match
        if h_lower in _COMMON_ALIASES:
            result[header] = _COMMON_ALIASES[h_lower]
            continue

        # 4. Fuzzy: check if header is substring of any label or vice versa
        best_match = None
        best_len = 0
        for label, fname in label_map.items():
            if h_lower in label or label in h_lower:
                if len(label) > best_len:
                    best_len = len(label)
                    best_match = fname
        if best_match:
            result[header] = best_match
            continue

    return result


def _resolve_value(client, field_name: str, value: Any, fields_def: dict) -> Any:
    if field_name not in fields_def:
        return value
    fdef = fields_def[field_name]
    if fdef.get("type") == "many2one" and isinstance(value, str) and value.strip():
        relation = fdef.get("relation", "")
        if relation:
            records = client.search_read(
                relation,
                domain=[("name", "=", value.strip())],
                fields=["id", "name"],
                limit=1,
            )
            if records:
                return records[0]["id"]
            # Try ilike
            records = client.search_read(
                relation,
                domain=[("name", "ilike", value.strip())],
                fields=["id", "name"],
                limit=1,
            )
            if records:
                return records[0]["id"]
            # No match found: return Odoo's "empty" for many2one, not the raw string
            return False
    return value


def _import_records(client, rows: list[dict[str, str]], model: str,
                    column_map: Optional[dict[str, str]], dry_run: bool) -> dict:
    if column_map is None:
        headers = list(rows[0].keys()) if rows else []
        column_map = map_columns(headers, model, client=client)

    fields_def = client.fields_get(model)

    created_count = 0
    skipped_count = 0
    errors: list[str] = []

    for row in rows:
        # Skip empty rows (all values blank)
        vals: dict[str, Any] = {}
        has_data = False
        for header, value in row.items():
            if header not in column_map:
                continue
            field_name = column_map[header]
            if value is not None and str(value).strip():
                has_data = True
                resolved = _resolve_value(client, field_name, value, fields_def)
                vals[field_name] = resolved

        if not has_data:
            skipped_count += 1
            continue

        if dry_run:
            continue

        try:
            client.create(model, vals)
            created_count += 1
        except (OdooClawError, ValueError, KeyError) as e:
            errors.append(f"Row {vals}: {e}")

    return {
        "created_count": created_count,
        "skipped_count": skipped_count,
        "errors": errors,
    }


def import_csv(client, filepath: str, model: Optional[str] = None,
               column_map: Optional[dict[str, str]] = None,
               dry_run: bool = False, registry=None) -> dict:
    """Import records from a CSV file into Odoo.

    Args:
        client: OdooClient instance.
        filepath: Path to the CSV file.
        model: Target model (auto-detected from headers if None).
        column_map: Optional explicit header-to-field mapping.
        dry_run: If True, validate without creating records.
        registry: Optional ModelRegistry for detecting custom models.

    Returns:
        Dict with created_count, skipped_count, and errors list.
    """
    filepath = _validate_filepath(filepath, "read")
    rows: list[dict[str, str]] = []
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        if model is None:
            model = detect_model(headers, registry=registry)
            if model is None:
                return {"created_count": 0, "skipped_count": 0,
                        "errors": ["Could not auto-detect model from headers"]}
        for row in reader:
            rows.append(row)

    return _import_records(client, rows, model, column_map, dry_run)


def import_excel(client, filepath: str, sheet: Optional[str] = None,
                 model: Optional[str] = None,
                 column_map: Optional[dict[str, str]] = None,
                 dry_run: bool = False, registry=None) -> dict:
    """Import records from an Excel file into Odoo.

    Args:
        client: OdooClient instance.
        filepath: Path to the .xlsx file.
        sheet: Optional sheet name (defaults to active sheet).
        model: Target model (auto-detected from headers if None).
        column_map: Optional explicit header-to-field mapping.
        dry_run: If True, validate without creating records.
        registry: Optional ModelRegistry for detecting custom models.

    Returns:
        Dict with created_count, skipped_count, and errors list.
    """
    filepath = _validate_filepath(filepath, "read")
    import openpyxl

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet:
        ws = wb[sheet]
    else:
        ws = wb.active

    row_iter = ws.iter_rows(values_only=True)
    headers = [str(c) if c is not None else "" for c in next(row_iter)]

    if model is None:
        model = detect_model(headers, registry=registry)
        if model is None:
            wb.close()
            return {"created_count": 0, "skipped_count": 0,
                    "errors": ["Could not auto-detect model from headers"]}

    rows: list[dict[str, str]] = []
    for row_values in row_iter:
        row_dict = {}
        for i, val in enumerate(row_values):
            if i < len(headers):
                row_dict[headers[i]] = str(val) if val is not None else ""
        rows.append(row_dict)

    wb.close()
    return _import_records(client, rows, model, column_map, dry_run)


def _format_value(value: Any) -> Any:
    if isinstance(value, (list, tuple)) and len(value) == 2:
        # Many2one: [id, name] -> name
        return value[1]
    if value is False:
        return ""
    return value


def export_records(client, model: str, domain: Optional[list] = None,
                   fields: Optional[list] = None, output_format: str = "csv",
                   filepath: Optional[str] = None) -> str:
    """Export Odoo records to a CSV or Excel file.

    Args:
        client: OdooClient instance.
        model: Odoo model to export.
        domain: Optional domain filter.
        fields: Fields to include (defaults to all non-id fields).
        output_format: Output format ('csv' or 'excel').
        filepath: Output path (auto-generated if None).

    Returns:
        Path to the created export file.
    """
    records = client.search_read(model, domain=domain or [], fields=fields or [], limit=0)

    if not fields and records:
        fields = [k for k in records[0].keys() if k != "id"]
    fields = fields or []

    if filepath is None:
        ext = ".xlsx" if output_format == "excel" else ".csv"
        filepath = os.path.join(_default_export_dir(), f"{model.replace('.', '_')}_export{ext}")
    else:
        filepath = _validate_filepath(filepath, "write")

    if output_format == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(fields)
        for rec in records:
            ws.append([_format_value(rec.get(f, "")) for f in fields])
        wb.save(filepath)
    else:
        with open(filepath, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(fields)
            for rec in records:
                writer.writerow([_format_value(rec.get(f, "")) for f in fields])

    return filepath


def generate_template(client, model: str, output_format: str = "csv",
                      filepath: Optional[str] = None) -> str:
    """Generate a blank CSV or Excel import template for a model.

    Args:
        client: OdooClient instance.
        model: Odoo model name.
        output_format: Output format ('csv' or 'excel').
        filepath: Output path (auto-generated if None).

    Returns:
        Path to the created template file.
    """
    fields_def = client.fields_get(model)

    headers = []
    for fname, fdef in sorted(fields_def.items()):
        if fname in _EXCLUDE_FROM_TEMPLATE:
            continue
        if fdef.get("type", "") in _EXCLUDE_TYPES_TEMPLATE:
            continue
        if fdef.get("readonly", False):
            continue
        headers.append(fname)

    if filepath is None:
        ext = ".xlsx" if output_format == "excel" else ".csv"
        filepath = os.path.join(_default_export_dir(), f"{model.replace('.', '_')}_template{ext}")
    else:
        filepath = _validate_filepath(filepath, "write")

    if output_format == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(filepath)
    else:
        with open(filepath, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(headers)

    return filepath
